"""
Build corporate_action_review.xlsx from corporate_action_suspects_clean.csv
Run: python build_review_xlsx.py
"""

import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

BASE = os.path.dirname(__file__)
SRC  = os.path.join(BASE, "corporate_action_suspects_clean.csv")
DST  = os.path.join(BASE, "corporate_action_review.xlsx")

# ── fill colours ────────────────────────────────────────────────────────────
FILLS = {
    "DROP_50":    PatternFill("solid", start_color="C6EFCE"),  # green
    "DROP_33":    PatternFill("solid", start_color="FFEB9C"),  # yellow
    "DROP_25":    PatternFill("solid", start_color="F4B942"),  # orange
    "DROP_OTHER": PatternFill("solid", start_color="FFFFFF"),  # no fill
}

HEADER_FILL = PatternFill("solid", start_color="1F4E79")  # dark blue header
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
CELL_FONT   = Font(name="Arial", size=10)

# ── sort order ──────────────────────────────────────────────────────────────
MAG_ORDER = {"DROP_50": 0, "DROP_33": 1, "DROP_25": 2, "DROP_OTHER": 3}

# ── load & sort ─────────────────────────────────────────────────────────────
with open(SRC, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    source_fields = reader.fieldnames
    rows = list(reader)

rows.sort(key=lambda r: (
    MAG_ORDER.get(r["magnitude_category"], 99),
    r["symbol"],
    r["date"],
))

fields = source_fields + ["confirmed_action"]

# ── build workbook ───────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Corporate Action Review"

# headers
for col_idx, name in enumerate(fields, 1):
    cell = ws.cell(row=1, column=col_idx, value=name)
    cell.font   = HEADER_FONT
    cell.fill   = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")

# data rows
for row_idx, row in enumerate(rows, 2):
    mag = row.get("magnitude_category", "DROP_OTHER")
    fill = FILLS.get(mag, FILLS["DROP_OTHER"])

    for col_idx, field in enumerate(fields, 1):
        value = row.get(field, "")
        # numeric columns
        if field in ("close_before", "close_after", "pct_change",
                     "peer_avg_change", "peers_used"):
            try:
                value = float(value)
            except (ValueError, TypeError):
                pass
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font      = CELL_FONT
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # confirmed_action stays empty — user fills manually

# ── freeze top row ───────────────────────────────────────────────────────────
ws.freeze_panes = "A2"

# ── auto-filter ──────────────────────────────────────────────────────────────
last_col = get_column_letter(len(fields))
ws.auto_filter.ref = f"A1:{last_col}1"

# ── auto-fit column widths ───────────────────────────────────────────────────
for col_idx, field in enumerate(fields, 1):
    col_letter = get_column_letter(col_idx)
    max_len = len(field)
    for row_idx in range(2, min(len(rows) + 2, 500)):  # sample first 498 rows
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is not None:
            max_len = max(max_len, len(str(val)))
    ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

# confirmed_action column — give it a bit more room
ws.column_dimensions[get_column_letter(len(fields))].width = 30

# row height for header
ws.row_dimensions[1].height = 20

wb.save(DST)

# ── summary ──────────────────────────────────────────────────────────────────
from collections import Counter
mag_counts = Counter(r["magnitude_category"] for r in rows)

print(f"Rows written: {len(rows):,}")
print(f"Saved:        {DST}")
print("\nBreakdown by magnitude_category:")
for cat in ["DROP_50", "DROP_33", "DROP_25", "DROP_OTHER"]:
    print(f"  {cat:<15} {mag_counts.get(cat, 0):>4}")
